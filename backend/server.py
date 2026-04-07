from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime
import openpyxl

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'ask_kitchen')]

# Create the main app
app = FastAPI(title="Aleppo Syrian Kitchen API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# ============== MODELS ==============

class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    cat_id: str  # Short ID like "Keb", "Yog"
    cover_image: str
    name_ar: str
    name_en: str
    name_sv: str

class Recipe(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    recipe_id: str  # Original ID from Excel
    category_id: str  # Links to Category.cat_id
    # Arabic
    name_ar: str
    time_ar: Optional[str] = None
    ingredients_ar: Optional[str] = None
    instructions_ar: Optional[str] = None
    servings_ar: Optional[str] = None
    decoration_ar: Optional[str] = None
    secrets_ar: Optional[str] = None
    pro_tips_ar: Optional[str] = None
    # English
    name_en: str
    time_en: Optional[str] = None
    ingredients_en: Optional[str] = None
    instructions_en: Optional[str] = None
    servings_en: Optional[str] = None
    decoration_en: Optional[str] = None
    secrets_en: Optional[str] = None
    pro_tips_en: Optional[str] = None
    # Swedish
    name_sv: str
    time_sv: Optional[str] = None
    ingredients_sv: Optional[str] = None
    instructions_sv: Optional[str] = None
    servings_sv: Optional[str] = None
    decoration_sv: Optional[str] = None
    secrets_sv: Optional[str] = None
    pro_tips_sv: Optional[str] = None
    # Common
    image: Optional[str] = None
    video_link: Optional[str] = None

class Feedback(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    email: str
    message: str
    timestamp: datetime = Field(default_factory=datetime.utcnow)

class FeedbackCreate(BaseModel):
    name: str
    email: str
    message: str

class AboutInfo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title_ar: str
    slogan_ar: str
    about_ar: str
    title_en: Optional[str] = None
    slogan_en: Optional[str] = None
    about_en: Optional[str] = None
    title_sv: Optional[str] = None
    slogan_sv: Optional[str] = None
    about_sv: Optional[str] = None

# ============== HELPER FUNCTIONS ==============

# Category name to ID mapping
CATEGORY_MAP = {
    'الكبب الحلبية': 'Keb',
    'أكلات باللبن الزبادي': 'Yog',
    'أكلات الحبوب': 'Grn',
    'أكلات متنوعة': 'Ass',
    'أكلات نباتية': 'Oil',
    'الخضار': 'Veg',
    'اللحوم': 'Met',
    'المحاشي': 'Stf',
    'المقبلات': 'App',
    'السلطات': 'Sld',
    'المعجنات': 'Pat',
}

# Google Drive folder ID from the shared link
GDRIVE_FOLDER_ID = "1s45Hum24BsF1OdQxCbeGugI8sNxTeR_D"

def clean_value(val):
    """Clean cell value, handling None and errors"""
    if val is None:
        return None
    val_str = str(val).strip()
    if val_str in ['#VALUE!', '#REF!', '#N/A', 'None', '']:
        return None
    return val_str

def get_image_url(image_path):
    """Convert image path to a usable URL"""
    if not image_path:
        return None
    # Clean the path
    path = clean_value(image_path)
    if not path:
        return None
    # Remove 'images/' or 'Images/' prefix and get filename
    filename = path.replace('images/', '').replace('Images/', '').replace('bilder/', '')
    # For now, return a placeholder or the filename
    # The images will be served from Google Drive or local storage
    return filename

def get_category_id(category_name_ar):
    """Get category ID from Arabic category name"""
    if not category_name_ar:
        return None
    name = category_name_ar.strip()
    return CATEGORY_MAP.get(name, 'Ass')  # Default to 'Various' if not found

# ============== API ROUTES ==============

@api_router.get("/")
async def root():
    return {"message": "Welcome to Aleppo Syrian Kitchen API", "version": "1.0"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}

# Categories
@api_router.get("/categories", response_model=List[Category])
async def get_categories():
    """Get all recipe categories"""
    categories = await db.categories.find().to_list(100)
    return [Category(**cat) for cat in categories]

@api_router.get("/categories/{cat_id}", response_model=Category)
async def get_category(cat_id: str):
    """Get a single category by ID"""
    category = await db.categories.find_one({"cat_id": cat_id})
    if not category:
        raise HTTPException(status_code=404, detail="Category not found")
    return Category(**category)

# Recipes
@api_router.get("/recipes", response_model=List[Recipe])
async def get_recipes(category_id: Optional[str] = None):
    """Get all recipes, optionally filtered by category"""
    query = {}
    if category_id:
        query["category_id"] = category_id
    recipes = await db.recipes.find(query).to_list(1000)
    return [Recipe(**recipe) for recipe in recipes]

@api_router.get("/recipes/{recipe_id}", response_model=Recipe)
async def get_recipe(recipe_id: str):
    """Get a single recipe by ID"""
    recipe = await db.recipes.find_one({"id": recipe_id})
    if not recipe:
        # Try by recipe_id (original Excel ID)
        recipe = await db.recipes.find_one({"recipe_id": recipe_id})
    if not recipe:
        raise HTTPException(status_code=404, detail="Recipe not found")
    return Recipe(**recipe)

# Feedback
@api_router.post("/feedback", response_model=Feedback)
async def create_feedback(feedback_input: FeedbackCreate):
    """Submit feedback/question"""
    feedback = Feedback(**feedback_input.dict())
    await db.feedback.insert_one(feedback.dict())
    return feedback

@api_router.get("/feedback", response_model=List[Feedback])
async def get_all_feedback():
    """Get all feedback (admin)"""
    feedback_list = await db.feedback.find().sort("timestamp", -1).to_list(1000)
    return [Feedback(**fb) for fb in feedback_list]

# About
@api_router.get("/about", response_model=AboutInfo)
async def get_about():
    """Get about information"""
    about = await db.about.find_one()
    if not about:
        # Return default
        return AboutInfo(
            title_ar="المطبخ الحلبي السوري",
            slogan_ar="أصالة الطعم من التراث الحلبي تحت شعار الأكل المهدا للملوك يتودا هكذا قالت جداتنا",
            about_ar="""نبذة عن صانعة الأطباق والتطبيق: 
أنا أم سامر سيدة حلبية سورية أصيلة، أصولي تعود الى 300 سنة في حلب حيث نشأت في بيئة تحب اللقمة الذكية الطيبة المطبوخة بحب وعناية لكي يوزع هذا الحب على العائلة والمحيطين بها. 

أنا من مواليد 1960 وقضيت 50 عاما في حب الطبخ ولم يكن مهنتي ولكنني عاشقة لتحضير الطعام للعائلة بكل أصنافة، وبما أنني في سنوات آخر العمر خطر ببالي أن أنقل كل ماأملك من معلومات تجعل المجتمع باسره عربي أو يتحدث الإنجليزية أو السويدية بحكم اقامتي في السويد يتلذذ بكل لقمة يعدها من وصفاتي التي أعتبرها ساحرة بحكم من حولي ومن تذوق الطعام على مائدتي المتواضعة.

سأقدم لكم في هذا التطبيق الوصفات الحلبية السورية الأصيلة وبعض الوصفات الخليجية بحكم اقامتي الطويلة في الخليج العربي أيضا فلا تستغربوا وجود وصفات خليجية لأنها أيضا أسرتني عندما كنت في الامارات وشعرت كم الثقافات تنتج طعاما يغير وجه العالم ويبعدنا عن الملل من تكرار مانأكل.

أم سامر – المطبخ الحلبي السوري""",
            title_en="Aleppo Syrian Kitchen",
            slogan_en="Slow cooked food fit for kings, as our grandmothers said",
            about_en="""About the dish and app creator:
I am Umm Samer, an authentic Syrian lady from Aleppo. My roots go back 300 years in Aleppo, where I grew up in an environment that loves delicious, lovingly prepared food to share with family and friends.

Born in 1960, I have spent 50 years loving cooking. It was never my profession, but I am passionate about preparing food for my family in all its varieties. In my later years, I decided to share all my knowledge with the community - whether Arabic, English, or Swedish speaking - so they can enjoy every bite prepared from my recipes that I consider magical.

In this app, I will present authentic Aleppo Syrian recipes and some Gulf recipes from my long stay in the Arabian Gulf. Don't be surprised to find Gulf recipes - they also captivated me when I was in the Emirates, and I realized how cultures produce food that changes the world.

Umm Samer - Aleppo Syrian Kitchen""",
            title_sv="Aleppo Syriskt Kök",
            slogan_sv="Långsamt tillagad mat som passar kungar, som våra mormödrar sa",
            about_sv="""Om skaparen av rätterna och appen:
Jag är Umm Samer, en äkta syrisk dam från Aleppo. Mina rötter går 300 år tillbaka i Aleppo, där jag växte upp i en miljö som älskar utsökt, kärleksfullt tillagad mat att dela med familj och vänner.

Född 1960, har jag spenderat 50 år med att älska matlagning. Det var aldrig mitt yrke, men jag brinner för att förbereda mat åt min familj i alla dess varianter. På mina äldre dagar bestämde jag mig för att dela all min kunskap med samhället - oavsett om de talar arabiska, engelska eller svenska.

I denna app presenterar jag autentiska syriska recept från Aleppo och några gulfrecept från min långa vistelse i Arabiska Gulfen.

Umm Samer - Aleppo Syriskt Kök"""
        )
    return AboutInfo(**about)

# Contact info
@api_router.get("/contact")
async def get_contact():
    """Get contact information"""
    return {
        "email": "askmalmo@gmail.com",
        "app_name": "Aleppo Syrian Kitchen",
        "app_name_ar": "المطبخ الحلبي السوري"
    }

# ============== DATA SEEDING ==============

@api_router.post("/seed")
async def seed_database():
    """Seed database from Excel files"""
    try:
        # Clear existing data
        await db.categories.delete_many({})
        await db.recipes.delete_many({})
        await db.about.delete_many({})
        
        # Seed categories
        categories_file = ROOT_DIR / 'categories_data.xlsx'
        if categories_file.exists():
            wb = openpyxl.load_workbook(categories_file)
            sheet = wb.active
            
            categories_added = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Has Cat_Id
                    cat = Category(
                        cat_id=clean_value(row[0]) or "",
                        cover_image=get_image_url(row[1]) or "",
                        name_en=clean_value(row[2]) or "",
                        name_sv=clean_value(row[3]) or "",
                        name_ar=clean_value(row[4]) or ""
                    )
                    await db.categories.insert_one(cat.dict())
                    categories_added += 1
            wb.close()
            logger.info(f"Added {categories_added} categories")
        
        # Seed recipes
        recipes_file = ROOT_DIR / 'recipes_data.xlsx'
        if recipes_file.exists():
            wb = openpyxl.load_workbook(recipes_file, data_only=True)
            sheet = wb['Ask Data']
            
            recipes_added = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0] and row[2]:  # Has ID and Name
                    recipe = Recipe(
                        recipe_id=str(clean_value(row[0]) or ""),
                        category_id=get_category_id(clean_value(row[1])),
                        # Arabic (columns 0-13)
                        name_ar=clean_value(row[2]) or "",
                        time_ar=clean_value(row[3]),
                        ingredients_ar=clean_value(row[4]),
                        instructions_ar=clean_value(row[5]),
                        servings_ar=clean_value(row[6]),
                        decoration_ar=clean_value(row[7]),
                        secrets_ar=clean_value(row[8]),
                        pro_tips_ar=clean_value(row[12]) if len(row) > 12 else None,
                        # English (columns 14-27)
                        name_en=clean_value(row[16]) or clean_value(row[2]) or "",
                        time_en=clean_value(row[17]),
                        ingredients_en=clean_value(row[18]),
                        instructions_en=clean_value(row[19]),
                        servings_en=clean_value(row[20]),
                        decoration_en=clean_value(row[21]),
                        secrets_en=clean_value(row[22]),
                        pro_tips_en=clean_value(row[26]) if len(row) > 26 else None,
                        # Swedish (columns 28-41)
                        name_sv=clean_value(row[30]) or clean_value(row[16]) or clean_value(row[2]) or "",
                        time_sv=clean_value(row[31]),
                        ingredients_sv=clean_value(row[32]),
                        instructions_sv=clean_value(row[33]),
                        servings_sv=clean_value(row[34]),
                        decoration_sv=clean_value(row[35]),
                        secrets_sv=clean_value(row[36]),
                        pro_tips_sv=clean_value(row[40]) if len(row) > 40 else None,
                        # Image
                        image=get_image_url(row[9])
                    )
                    await db.recipes.insert_one(recipe.dict())
                    recipes_added += 1
            wb.close()
            logger.info(f"Added {recipes_added} recipes")
        
        # Seed about info
        about = AboutInfo(
            title_ar="المطبخ الحلبي السوري",
            slogan_ar="الأكل المهدا للملوك يتودا هكذا قالت جداتنا",
            about_ar="""نبذة عن صانعة الأطباق والتطبيق: 
أنا أم سامر سيدة حلبية سورية أصيلة، أصولي تعود الى 300 سنة في حلب حيث نشأت في بيئة تحب اللقمة الذكية الطيبة المطبوخة بحب وعناية لكي يوزع هذا الحب على العائلة والمحيطين بها.

أنا من مواليد 1960 وقضيت 50 عاما في حب الطبخ ولم يكن مهنتي ولكنني عاشقة لتحضير الطعام للعائلة بكل أصنافة.""",
            title_en="Aleppo Syrian Kitchen",
            slogan_en="Food fit for kings, as our grandmothers said",
            about_en="""About the dish and app creator:
I am Umm Samer, an authentic Syrian lady from Aleppo. My roots go back 300 years in Aleppo, where I grew up in an environment that loves delicious, lovingly prepared food to share with family and friends.

Born in 1960, I have spent 50 years loving cooking. It was never my profession, but I am passionate about preparing food for my family in all its varieties.""",
            title_sv="Aleppo Syriskt Kök",
            slogan_sv="Mat som passar kungar, som våra mormödrar sa",
            about_sv="""Om skaparen av rätterna och appen:
Jag är Umm Samer, en äkta syrisk dam från Aleppo. Mina rötter går 300 år tillbaka i Aleppo, där jag växte upp i en miljö som älskar utsökt, kärleksfullt tillagad mat att dela med familj och vänner.

Född 1960, har jag spenderat 50 år med att älska matlagning. Det var aldrig mitt yrke, men jag brinner för att förbereda mat åt min familj i alla dess varianter."""
        )
        await db.about.insert_one(about.dict())
        
        return {
            "success": True,
            "message": "Database seeded successfully",
            "categories": categories_added if 'categories_added' in dir() else 0,
            "recipes": recipes_added if 'recipes_added' in dir() else 0
        }
    except Exception as e:
        logger.error(f"Error seeding database: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/stats")
async def get_stats():
    """Get database statistics"""
    categories_count = await db.categories.count_documents({})
    recipes_count = await db.recipes.count_documents({})
    feedback_count = await db.feedback.count_documents({})
    return {
        "categories": categories_count,
        "recipes": recipes_count,
        "feedback": feedback_count
    }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
